from playwright.sync_api import sync_playwright
import time
import os
import argparse
import re
from pathlib import Path
import sys
import openpyxl
from openpyxl.cell.cell import MergedCell

# --- CONFIGURATION ---
# Get the root directory of the project
ROOT_DIR = Path(__file__).resolve().parent

# Define the expected Excel file name for the student
DEFAULT_EXCEL_CANDIDATES = [
    str(ROOT_DIR / "IT23320932_Test cases.xlsx"), 
]

# Sheet name within the Excel file
DEFAULT_SHEET_NAME = " Test cases" 

# Target website for transliteration testing[cite: 1]
DEFAULT_FRONTEND_URL = "https://www.pixelssuite.com/chat-translator" 

# Mapping column headers for identification[cite: 2, 3]
DEFAULT_INPUT_COLUMN_CANDIDATES = ["Input", "Singlish"]
DEFAULT_EXPECTED_COLUMN_CANDIDATES = ["Expected output", "Expected"]
DEFAULT_ACTUAL_COLUMN_CANDIDATES = ["Actual output"]
DEFAULT_STATUS_COLUMN_CANDIDATES = ["Status"]

# Automation and synchronization settings
DEFAULT_WAIT_MS = 5000
DEFAULT_RETRIES = 5
DEFAULT_RETRY_WAIT_MS = 1000
DEFAULT_TYPE_DELAY_MS = 30
DEFAULT_TIMEOUT_MS = 60000

def _configure_stdout():
    """Ensure console output handles Unicode (Sinhala) characters correctly[cite: 3]."""
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass

def _resolve_path(p: str | None) -> str | None:
    """Resolve file paths relative to the project root[cite: 3]."""
    if not p: return None
    path = Path(p)
    if path.is_absolute(): return str(path)
    return str((ROOT_DIR / path).resolve())

def _normalize_header(value) -> str:
    """Standardize header strings for better matching[cite: 3]."""
    if value is None: return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())

def _find_header_row(ws, max_scan_rows: int) -> int:
    """Scan the Excel sheet to find the header row based on keywords[cite: 3]."""
    for r in range(1, max_scan_rows + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        norms = {_normalize_header(v) for v in values if v}
        if "input" in norms or "singlish" in norms:
            return r
    return 1

def _merged_top_left_cell(ws, row: int, col: int):
    """Handle merged cells in Excel to ensure value updates are applied correctly[cite: 3]."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell): return cell
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return cell

def _set_cell_value(ws, row: int, col: int, value):
    """Update a specific cell value while considering merged cell structures[cite: 3]."""
    cell = _merged_top_left_cell(ws, row, col)
    cell.value = value

def _find_column_index(ws, header_row: int, candidates: list[str]) -> int | None:
    """Identify the column index based on potential header candidates[cite: 3]."""
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val and _normalize_header(val) in [_normalize_header(cand) for cand in candidates]:
            return c
    return None

def _read_output(page, output_locator) -> str:
    """Extract output text from the website's results area[cite: 3]."""
    try:
        return str(output_locator.input_value() or output_locator.inner_text() or "").strip()
    except:
        return ""

def run_test():
    """Main execution function for automation[cite: 3]."""
    _configure_stdout()
    
    # Locate the Excel file[cite: 3]
    excel_path = None
    for cand in DEFAULT_EXCEL_CANDIDATES:
        if os.path.exists(cand):
            excel_path = cand
            break
    
    if not excel_path:
        print(f"Error: Excel file not found. Checked: {DEFAULT_EXCEL_CANDIDATES}")
        return

    # Load Workbook and Sheet[cite: 3]
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[DEFAULT_SHEET_NAME] if DEFAULT_SHEET_NAME in wb.sheetnames else wb.active
    
    # Map headers to indices[cite: 3]
    header_row = _find_header_row(ws, 20)
    input_col = _find_column_index(ws, header_row, DEFAULT_INPUT_COLUMN_CANDIDATES)
    expected_col = _find_column_index(ws, header_row, DEFAULT_EXPECTED_COLUMN_CANDIDATES)
    actual_col = _find_column_index(ws, header_row, DEFAULT_ACTUAL_COLUMN_CANDIDATES) or (ws.max_column + 1)
    status_col = _find_column_index(ws, header_row, DEFAULT_STATUS_COLUMN_CANDIDATES) or (ws.max_column + 2)

    if not input_col:
        print("Error: Could not identify Input column in the provided Excel sheet.")
        return

    with sync_playwright() as p:
        # Launch the browser (Headed mode for visibility)[cite: 3]
        browser = p.chromium.launch(headless=False) 
        page = browser.new_page()
        print(f"Navigating to: {DEFAULT_FRONTEND_URL}")
        page.goto(DEFAULT_FRONTEND_URL)
        
        # UI Element Locators[cite: 1, 3]
        input_locator = page.locator('textarea[placeholder*="English"]').first
        output_locator = page.locator('textarea[placeholder*="Sinhala"]').first
        btn_locator = page.get_by_role("button", name=re.compile(r"Transliterate", re.IGNORECASE)).first

        # Process each row[cite: 3]
        for row in range(header_row + 1, ws.max_row + 1):
            singlish_val = ws.cell(row=row, column=input_col).value
            if not singlish_val: continue
            
            expected_val = ws.cell(row=row, column=expected_col).value if expected_col else ""
            
            print(f"Testing [Row {row}]: {singlish_val}")
            
            # Browser interactions[cite: 3]
            input_locator.fill("") # Clear input
            input_locator.fill(str(singlish_val)) # Enter Singlish text
            btn_locator.click() # Click transliterate button
            
            # Wait for output generation[cite: 3]
            page.wait_for_timeout(DEFAULT_WAIT_MS)
            
            actual_sinhala = _read_output(page, output_locator)
            _set_cell_value(ws, row, actual_col, actual_sinhala)
            
            # Perform strict comparison for status reporting[cite: 3]
            status = "PASS" if str(actual_sinhala).strip() == str(expected_val or "").strip() else "FAIL"
            _set_cell_value(ws, row, status_col, status)
            print(f" -> Result: {status}")

        # Save the updated Excel report[cite: 3]
        wb.save(excel_path)
        print(f"\nAutomation Completed. Results saved to: {excel_path}")
        browser.close()

if __name__ == "__main__":
    run_test()